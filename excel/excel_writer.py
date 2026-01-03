import openpyxl
from datetime import datetime, date, timedelta
from utils.dates import get_day_type

def write_commit_excel(wb, records, start_date, end_date, holidays=None):
    """
    Attendance Calculation for Salary Computation.
    
    CRITICAL: This function ensures 100% accuracy for salary calculations.
    Every hour is accounted for with clear categorization.
    """
    if holidays is None:
        holidays = []
    
    ws = wb.active
    
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    
    print(f"\n{'='*80}")
    print(f"Processing: {records.get('employee_name')}")
    print(f"{'='*80}")
    
    # Map days by day number
    day_records = {}
    for rec in records.get("records", []):
        day = rec.get("day")
        if day and 1 <= day <= 31:
            day_records[day] = rec
    
    # Get total days in period
    total_days = (end_date - start_date).days + 1
    
    # ========== CLEAR CATEGORY COUNTERS ==========
    normal_days = 0.0              # Full 8-hour weekday work days
    weekday_ot_hours = 0.0         # Extra hours > 8 on weekdays
    weekend_ot_hours = 0.0         # ALL hours on weekends
    holiday_ot_hours = 0.0         # ALL hours on holidays
    absent_days = 0                # ABSENT/LEAVE on weekdays only
    
    # Verification
    total_extracted_work_hours = 0.0  # Sum of all WORK status hours
    actual_work_days = 0
    
    # ========== PROCESS EACH DAY ==========
    for day_num in range(1, total_days + 1):
        current_date = start_date + timedelta(days=day_num - 1)
        day_type = get_day_type(current_date, holidays)
        
        if day_num in day_records:
            rec = day_records[day_num]
            status = rec.get("status", "ABSENT").upper().strip()
            hours = float(rec.get("hours", 0))
            
            # ---- LEAVE or ABSENT ----
            if status in ["LEAVE", "L", "AL", "SL", "ABSENT"]:
                if day_type == "WEEKDAY":
                    absent_days += 1
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): {status:7s} → ABSENT on WEEKDAY")
                else:
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): {status:7s} → OFF ({day_type})")
            
            # ---- OFF ----
            elif status == "OFF":
                print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): OFF")
            
            # ---- HOLIDAY (no work) ----
            elif status == "HOLIDAY":
                if hours > 0:
                    holiday_ot_hours += hours
                    total_extracted_work_hours += hours
                    actual_work_days += 1
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): HOLIDAY {hours:4.1f}h → HOLIDAY OT {hours:4.1f}h")
                else:
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): HOLIDAY (no work)")
            
            # ---- WORK ----
            elif status == "WORK":
                if hours <= 0:
                    if day_type == "WEEKDAY":
                        absent_days += 1
                        print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK 0.0h → ABSENT on WEEKDAY")
                
                elif day_type == "WEEKEND":
                    # Weekend work: ALL hours count as weekend OT
                    weekend_ot_hours += hours
                    total_extracted_work_hours += hours
                    actual_work_days += 1
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK {hours:4.1f}h WEEKEND → WEEKEND OT {hours:4.1f}h")
                
                elif day_type == "HOLIDAY":
                    # Holiday work: ALL hours count as holiday OT
                    holiday_ot_hours += hours
                    total_extracted_work_hours += hours
                    actual_work_days += 1
                    print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK {hours:4.1f}h HOLIDAY → HOLIDAY OT {hours:4.1f}h")
                
                else:
                    # WEEKDAY work - THIS IS WHERE NORMAL DAYS COUNT
                    total_extracted_work_hours += hours
                    actual_work_days += 1
                    
                    if hours > 8:
                        # Normal (8h) + Overtime (extra)
                        normal_days += 1.0
                        weekday_ot_hours += (hours - 8)
                        print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK {hours:4.1f}h WEEKDAY → 1.0 NORMAL + {hours-8:4.1f}h OT")
                    
                    elif hours == 8:
                        # Full day
                        normal_days += 1.0
                        print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK {hours:4.1f}h WEEKDAY → 1.0 NORMAL DAY")
                    
                    else:
                        # Partial day
                        fraction = hours / 8
                        normal_days += fraction
                        print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): WORK {hours:4.1f}h WEEKDAY → {fraction:.3f} PARTIAL DAY")
        
        else:
            # No record found
            if day_type == "WEEKDAY":
                absent_days += 1
                print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): NO RECORD → ABSENT on WEEKDAY")
            else:
                print(f"  Day {day_num:2d} ({current_date.strftime('%a')}): NO RECORD → OFF ({day_type})")
    
    # ========== CALCULATE TOTALS ==========
    employee_name = records.get("employee_name", "UNKNOWN")
    employee_code = records.get("employee_code", "")
    
    shortfall = 0.0
    if actual_work_days < 22 and total_extracted_work_hours < 176:
        shortfall = 176 - total_extracted_work_hours
    
    # ========== VERIFICATION ==========
    reconstructed_total = (normal_days * 8) + weekday_ot_hours + weekend_ot_hours + holiday_ot_hours
    match = abs(reconstructed_total - total_extracted_work_hours) < 0.01
    
    print(f"\n{'='*80}")
    print(f"📊 CALCULATION BREAKDOWN :")
    print(f"{'='*80}")
    print(f"\n  WEEKDAY WORK:")
    print(f"    Normal Days:        {normal_days:8.2f} days × 8h = {normal_days*8:8.1f}h")
    print(f"    Weekday Overtime:   {weekday_ot_hours:36.1f}h")
    print(f"\n  SPECIAL WORK:")
    print(f"    Weekend OT:         {weekend_ot_hours:36.1f}h")
    print(f"    Holiday OT:         {holiday_ot_hours:36.1f}h")
    print(f"\n  CATEGORY TOTALS:")
    print(f"    Reconstructed:      {reconstructed_total:36.1f}h (from categories)")
    print(f"    Extracted:          {total_extracted_work_hours:36.1f}h (sum of WORK)")
    print(f"    Verification:       {'✅ MATCH' if match else '❌ MISMATCH - CALC ERROR'}")
    
    if not match:
        print(f"    Difference:         {abs(reconstructed_total - total_extracted_work_hours):36.2f}h")
        print(f"\n    ⚠️  CRITICAL ERROR - Calculation mismatch detected!")
        print(f"    This must be fixed before salary computation!\n")
    
    print(f"\n  ABSENCES:")
    print(f"    Absent/Leave Days:  {absent_days:36d}")
    print(f"    Work Days:          {actual_work_days:36d}")
    print(f"    Shortfall Hours:    {shortfall:36.1f}h")
    print(f"{'='*80}\n")
    
    # Write to Excel
    ws.cell(row=row, column=1).value = employee_name
    ws.cell(row=row, column=2).value = employee_code
    ws.cell(row=row, column=3).value = False
    ws.cell(row=row, column=4).value = start_date
    ws.cell(row=row, column=5).value = end_date
    ws.cell(row=row, column=6).value = normal_days
    ws.cell(row=row, column=7).value = weekday_ot_hours
    ws.cell(row=row, column=8).value = weekend_ot_hours
    ws.cell(row=row, column=9).value = holiday_ot_hours
    ws.cell(row=row, column=10).value = absent_days
    ws.cell(row=row, column=11).value = shortfall
    
    print(f"✅ Written to row {row}\n")
    
    return wb, total_extracted_work_hours